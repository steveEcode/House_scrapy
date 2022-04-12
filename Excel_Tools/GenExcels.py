from openpyxl import Workbook

class Excel():

    def __init__(self,save_path):
        self.save_path = save_path

    def gen_excel(self,results):
        workbook = Workbook()
        sheet = workbook._sheets[0]

        row = 1
        col = 1
        title_list = []
        for r in results[0].keys():
            title_list.append(r)
        # title_list = ['zip_code','house_id','address','price','detailUrl','availabilityDate']
        for t in title_list:
            sheet.cell(row=row,column=col,value=t)
            col += 1

        row = 2
        for r in results:
            col = 1
            for i in r.values():
                sheet.cell(row=row,column=col,value=str(i))
                col += 1
            row += 1
        workbook.save(filename=self.save_path)